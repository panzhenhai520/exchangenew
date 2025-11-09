# -*- coding: utf-8 -*-
"""
BOT报表生成服务 V3 - 完整列映射修复版
确保所有字段都能正确填充
版本: v3.0
创建日期: 2025-10-08
"""

import os
import shutil
from datetime import datetime
from typing import Dict, Any, Optional
import logging

import openpyxl
from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


class BOTTemplateGeneratorV3:
    """BOT Excel报表生成器 V3 - 修复完整列映射"""

    TEMPLATE_PATH = r"D:\Code\ExchangeNew\Re\BOT( Save Excel, PDF, Bot).xlsx"

    @classmethod
    def generate_report(
        cls,
        db_session: Session,
        branch_id: int,
        report_month: int,
        report_year: int,
        output_path: Optional[str] = None
    ) -> bytes:
        """生成BOT报表"""
        try:
            if not os.path.exists(cls.TEMPLATE_PATH):
                raise FileNotFoundError(f"模板文件不存在: {cls.TEMPLATE_PATH}")

            # 复制模板
            temp_path = cls._create_temp_copy()
            logger.info(f"已复制模板到: {temp_path}")

            # 打开工作簿
            wb = openpyxl.load_workbook(temp_path)

            # 获取网点信息
            branch_info = cls._get_branch_info(db_session, branch_id)

            # 填充数据
            cls._fill_provider_info(wb, branch_info, report_month, report_year)
            cls._fill_buy_fx_data_v3(wb, db_session, branch_id, report_month, report_year)
            cls._fill_sell_fx_data_v3(wb, db_session, branch_id, report_month, report_year)
            cls._fill_fcd_data_v3(wb, db_session, branch_id, report_month, report_year)

            # 保存
            if output_path:
                wb.save(output_path)
                logger.info(f"BOT报表已保存: {output_path}")
                with open(output_path, 'rb') as f:
                    return f.read()
            else:
                import io
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return output.getvalue()

        except Exception as e:
            logger.error(f"生成BOT报表失败: {str(e)}")
            raise

        finally:
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass

    @classmethod
    def _create_temp_copy(cls) -> str:
        temp_dir = os.path.join(os.path.dirname(cls.TEMPLATE_PATH), 'temp')
        os.makedirs(temp_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        temp_path = os.path.join(temp_dir, f'BOT_temp_{timestamp}.xlsx')
        shutil.copy2(cls.TEMPLATE_PATH, temp_path)
        return temp_path

    @classmethod
    def _get_branch_info(cls, db_session: Session, branch_id: int) -> Dict[str, Any]:
        sql = text("""
            SELECT
                branch_name,
                address,
                phone_number,
                license_number,
                branch_code,
                company_full_name,
                bot_sender_code,
                bot_branch_area_code,
                bot_license_number
            FROM branches WHERE id = :branch_id
        """)

        result = db_session.execute(sql, {'branch_id': branch_id}).fetchone()

        if not result:
            return {
                'institution_code': '0105549142901',
                'license_holder_name': 'บริษัท วางสวิตต์ อินเตอร์เนชั่นแนล จำกัด',
                'license_no': 'MC325670003',
                'branch_name': 'บริษัท วางสวิตต์ อินเตอร์เนชั่นแนล จำกัด',
                'branch_area_code': '001'
            }

        branch_name, _, _, license_number, branch_code, company_full_name, bot_sender_code, bot_branch_area_code, bot_license_number = result

        institution_code = bot_sender_code or '0105549142901'
        license_holder_name = company_full_name or branch_name or 'บริษัท วางสวิตต์ อินเตอร์เนชั่นแนล จำกัด'
        license_no = bot_license_number or license_number or 'MC325670003'
        area_code = bot_branch_area_code or branch_code or str(branch_id).zfill(3)

        return {
            'institution_code': institution_code,
            'license_holder_name': license_holder_name,
            'license_no': license_no,
            'branch_name': branch_name or license_holder_name,
            'branch_area_code': area_code
        }

    @classmethod
    def _fill_provider_info(cls, wb, branch_info, report_month, report_year):
        ws = wb['Provider Info']

        ws['B2'] = branch_info['institution_code']
        ws['B3'] = branch_info['license_holder_name']
        ws['B4'] = branch_info['license_no']
        ws['B5'] = branch_info['branch_name']
        ws['B6'] = branch_info['branch_area_code']

        month_names_th = [
            '', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
            'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม'
        ]
        ws['B7'] = month_names_th[report_month] if 1 <= report_month <= 12 else ''
        ws['B8'] = report_year

        gregorian_year = report_year - 543
        import calendar
        last_day = calendar.monthrange(gregorian_year, report_month)[1]
        ws['B9'] = f"{gregorian_year}-{str(report_month).zfill(2)}-{str(last_day).zfill(2)}"

        logger.info(f"已填充Provider Info: {branch_info['branch_name']}, {report_year}/{report_month}")

    @classmethod
    def _fill_buy_fx_data_v3(cls, wb, db_session, branch_id, report_month, report_year):
        """填充Buy FX数据 - V3完整列映射"""
        ws = wb['Buy FX']
        gregorian_year = report_year - 543

        sql = text("""
            SELECT
                transaction_date, transaction_no,
                customer_id_type, customer_id_number, customer_name,
                customer_country_code, buy_currency_code, buy_amount,
                local_amount, exchange_rate, usd_equivalent, remarks
            FROM bot_buyfx
            WHERE branch_id = :branch_id
              AND YEAR(transaction_date) = :year
              AND MONTH(transaction_date) = :month
            ORDER BY transaction_date, transaction_no
        """)

        results = db_session.execute(sql, {
            'branch_id': branch_id,
            'year': gregorian_year,
            'month': report_month
        }).fetchall()

        current_row = 9  # 数据从第9行开始
        for idx, row_data in enumerate(results, 1):
            tx_date = row_data[0]

            # A列: 序号
            ws.cell(row=current_row, column=1, value=idx)

            # B列: 客户类型（ประเภทของลูกค้า）
            customer_type = cls._get_customer_type(row_data[2])  # 基于证件类型判断
            ws.cell(row=current_row, column=2, value=customer_type)

            # C列: 客户姓名
            ws.cell(row=current_row, column=3, value=row_data[4])  # customer_name

            # D列: 证件类型代码（公式会自动填充，但我们也可以手动设置）
            # 留空，让公式自动计算

            # E列: 证件号码
            ws.cell(row=current_row, column=5, value=row_data[3])  # customer_id_number

            # F列: 国籍代码
            country_code = row_data[5] or 'TH'
            ws.cell(row=current_row, column=6, value=country_code)

            # G列: 国籍描述（由公式自动填充）
            # 公式: =IFERROR(VLOOKUP(F9,Dimension!$G$3:$H$252,2,FALSE),"")

            # H列: 交易目的（由公式自动填充）
            # 公式: =IF(OR(B9=...),"เดินทาง/ท่องเที่ยว","")

            # I-J列: 交易地点和媒介（由公式自动填充）

            # K列: 货币代码
            currency_code = row_data[6] or 'USD'
            ws.cell(row=current_row, column=11, value=currency_code)

            # L列: 货币描述（由公式自动填充）
            # 公式: =IFERROR(VLOOKUP(K9,Dimension!$J$3:$K$179,2,FALSE),"")

            # M列: 汇率 ⭐ 重要
            ws.cell(row=current_row, column=13, value=float(row_data[9]) if row_data[9] else 0)

            # N列: 外币金额 ⭐ 重要
            ws.cell(row=current_row, column=14, value=float(row_data[7]) if row_data[7] else 0)

            # O-P列: 支付地点和媒介（由公式自动填充）

            # Q列: 泰铢金额（由公式自动计算 =ROUND(M*N,2)）⭐
            # 这列不需要手动填充，公式会自动计算

            # R列: 备注 ⭐
            ws.cell(row=current_row, column=18, value=row_data[11] or '')

            current_row += 1

        logger.info(f"已填充Buy FX数据: {len(results)} 条记录")

    @classmethod
    def _fill_sell_fx_data_v3(cls, wb, db_session, branch_id, report_month, report_year):
        """填充Sell FX数据 - V3完整列映射"""
        ws = wb['Sell FX']
        gregorian_year = report_year - 543

        sql = text("""
            SELECT
                transaction_date, transaction_no,
                customer_id_type, customer_id_number, customer_name,
                customer_country_code, sell_currency_code, sell_amount,
                local_amount, exchange_rate, usd_equivalent, remarks
            FROM bot_sellfx
            WHERE branch_id = :branch_id
              AND YEAR(transaction_date) = :year
              AND MONTH(transaction_date) = :month
            ORDER BY transaction_date, transaction_no
        """)

        results = db_session.execute(sql, {
            'branch_id': branch_id,
            'year': gregorian_year,
            'month': report_month
        }).fetchall()

        current_row = 9
        for idx, row_data in enumerate(results, 1):
            ws.cell(row=current_row, column=1, value=idx)  # A: 序号
            ws.cell(row=current_row, column=2, value=cls._get_customer_type(row_data[2]))  # B: 客户类型
            ws.cell(row=current_row, column=3, value=row_data[4])  # C: 客户姓名
            ws.cell(row=current_row, column=5, value=row_data[3])  # E: 证件号码
            ws.cell(row=current_row, column=6, value=row_data[5] or 'TH')  # F: 国籍代码
            ws.cell(row=current_row, column=11, value=row_data[6] or 'USD')  # K: 货币代码
            ws.cell(row=current_row, column=13, value=float(row_data[9]) if row_data[9] else 0)  # M: 汇率
            ws.cell(row=current_row, column=14, value=float(row_data[7]) if row_data[7] else 0)  # N: 外币金额
            ws.cell(row=current_row, column=18, value=row_data[11] or '')  # R: 备注

            current_row += 1

        logger.info(f"已填充Sell FX数据: {len(results)} 条记录")

    @classmethod
    def _fill_fcd_data_v3(cls, wb, db_session, branch_id, report_month, report_year):
        """填充FCD数据 - V3版本"""
        ws = wb['FCD']
        gregorian_year = report_year - 543

        sql = text("""
            SELECT
                account_open_date, bank_name, account_number,
                currency_code, balance, transaction_amount,
                usd_equivalent, remarks
            FROM bot_fcd
            WHERE branch_id = :branch_id
              AND YEAR(account_open_date) = :year
              AND MONTH(account_open_date) = :month
            ORDER BY account_open_date, account_number
        """)

        results = db_session.execute(sql, {
            'branch_id': branch_id,
            'year': gregorian_year,
            'month': report_month
        }).fetchall()

        current_row = 8  # FCD从第8行开始
        for idx, row_data in enumerate(results, 1):
            open_date = row_data[0]

            ws.cell(row=current_row, column=1, value=idx)  # A: 序号
            ws.cell(row=current_row, column=2, value=row_data[1] or '')  # B: 银行名称
            # C列: 银行代码（由公式自动填充）
            ws.cell(row=current_row, column=4, value=row_data[2] or '')  # D: 账号
            ws.cell(row=current_row, column=5, value=row_data[3] or '')  # E: 货币代码
            # F列: 货币描述（由公式自动填充）
            ws.cell(row=current_row, column=7, value=row_data[7] or '')  # G: 备注

            current_row += 1

        logger.info(f"已填充FCD数据: {len(results)} 条记录")

    @staticmethod
    def _get_customer_type(id_type: str) -> str:
        """根据证件类型判断客户类型"""
        type_mapping = {
            'Passport': 'ชาวต่างชาติ',  # 外国人
            'ID Card': 'คนไทย',  # 泰国人
            'Thai ID': 'คนไทย',  # 泰国人
            'National ID': 'คนไทย',  # 泰国人
            'Corporate': 'นิติบุคคลไทย'  # 泰国法人
        }
        return type_mapping.get(id_type, 'ชาวต่างชาติ')
