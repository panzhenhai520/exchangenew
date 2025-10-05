# -*- coding: utf-8 -*-
"""
BOT报告API路由
提供BOT报表查询和导出的核心接口
版本: v1.0
创建日期: 2025-10-02
"""

from flask import Blueprint, request, jsonify, g, send_file
from functools import wraps
from services.db_service import SessionLocal
from services.auth_service import token_required, permission_required
from sqlalchemy import text
import traceback
import json
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 创建Blueprint
app_bot = Blueprint('app_bot', __name__)


# 权限装饰器
def bot_permission_required(permission):
    """BOT权限检查装饰器"""
    return permission_required(permission)


@app_bot.route('/api/bot/t1-buy-fx', methods=['GET'])
@token_required
@bot_permission_required('bot_report_view')
def get_t1_buy_fx():
    """
    查询T+1买入外币报表数据

    GET /api/bot/t1-buy-fx?date=2025-10-01

    查询参数:
    - date: 交易日期，默认昨天

    响应:
    {
        "success": true,
        "data": {
            "date": "2025-10-01",
            "items": [...],
            "total_count": 10,
            "total_amount_thb": 5000000
        }
    }
    """
    session = SessionLocal()

    try:
        # 获取日期参数，默认昨天
        date_str = request.args.get('date')
        if not date_str:
            yesterday = datetime.now() - timedelta(days=1)
            date_str = yesterday.strftime('%Y-%m-%d')

        # 获取当前用户的branch_id
        branch_id = g.current_user.get('branch_id')

        # 查询交易记录
        sql = text("""
            SELECT
                t.id,
                t.transaction_no,
                t.customer_id,
                c.currency_code,
                c.currency_name,
                t.foreign_amount,
                t.local_amount,
                t.exchange_rate,
                t.transaction_time,
                t.operator_id,
                t.exchange_type,
                t.funding_source
            FROM exchange_transactions t
            LEFT JOIN currencies c ON t.currency_id = c.id
            WHERE t.branch_id = :branch_id
                AND DATE(t.transaction_time) = :date
                AND t.direction = 'buy'
                AND t.bot_flag = 1
                AND t.status = 'completed'
            ORDER BY t.transaction_time ASC
        """)

        result = session.execute(sql, {'branch_id': branch_id, 'date': date_str})
        items = [dict(row._mapping) for row in result]

        # 计算汇总
        total_count = len(items)
        total_amount_thb = sum([float(item['local_amount'] or 0) for item in items])

        return jsonify({
            'success': True,
            'data': {
                'date': date_str,
                'items': items,
                'total_count': total_count,
                'total_amount_thb': total_amount_thb
            }
        })

    except Exception as e:
        print(f"Error in get_t1_buy_fx: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'查询买入外币报表失败: {str(e)}'
        }), 500

    finally:
        session.close()


@app_bot.route('/api/bot/t1-sell-fx', methods=['GET'])
@token_required
@bot_permission_required('bot_report_view')
def get_t1_sell_fx():
    """
    查询T+1卖出外币报表数据

    GET /api/bot/t1-sell-fx?date=2025-10-01

    查询参数:
    - date: 交易日期，默认昨天

    响应:
    {
        "success": true,
        "data": {
            "date": "2025-10-01",
            "items": [...],
            "total_count": 8,
            "total_amount_thb": 3000000
        }
    }
    """
    session = SessionLocal()

    try:
        # 获取日期参数，默认昨天
        date_str = request.args.get('date')
        if not date_str:
            yesterday = datetime.now() - timedelta(days=1)
            date_str = yesterday.strftime('%Y-%m-%d')

        # 获取当前用户的branch_id
        branch_id = g.current_user.get('branch_id')

        # 查询交易记录
        sql = text("""
            SELECT
                t.id,
                t.transaction_no,
                t.customer_id,
                c.currency_code,
                c.currency_name,
                t.foreign_amount,
                t.local_amount,
                t.exchange_rate,
                t.transaction_time,
                t.operator_id,
                t.exchange_type,
                t.funding_source
            FROM exchange_transactions t
            LEFT JOIN currencies c ON t.currency_id = c.id
            WHERE t.branch_id = :branch_id
                AND DATE(t.transaction_time) = :date
                AND t.direction = 'sell'
                AND t.bot_flag = 1
                AND t.status = 'completed'
            ORDER BY t.transaction_time ASC
        """)

        result = session.execute(sql, {'branch_id': branch_id, 'date': date_str})
        items = [dict(row._mapping) for row in result]

        # 计算汇总
        total_count = len(items)
        total_amount_thb = sum([float(item['local_amount'] or 0) for item in items])

        return jsonify({
            'success': True,
            'data': {
                'date': date_str,
                'items': items,
                'total_count': total_count,
                'total_amount_thb': total_amount_thb
            }
        })

    except Exception as e:
        print(f"Error in get_t1_sell_fx: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'查询卖出外币报表失败: {str(e)}'
        }), 500

    finally:
        session.close()


@app_bot.route('/api/bot/export-buy-fx', methods=['GET'])
@token_required
@bot_permission_required('bot_report_export')
def export_buy_fx_excel():
    """
    导出买入外币报表Excel

    GET /api/bot/export-buy-fx?date=2025-10-01

    查询参数:
    - date: 交易日期，默认昨天

    响应:
    Excel文件下载
    """
    session = SessionLocal()

    try:
        # 获取日期参数
        date_str = request.args.get('date')
        if not date_str:
            yesterday = datetime.now() - timedelta(days=1)
            date_str = yesterday.strftime('%Y-%m-%d')

        # 获取当前用户的branch_id
        branch_id = g.current_user.get('branch_id')

        # 查询数据
        sql = text("""
            SELECT
                t.transaction_no as 交易编号,
                DATE_FORMAT(t.transaction_time, '%Y-%m-%d %H:%i:%s') as 交易时间,
                t.customer_id as 客户证件号,
                c.currency_code as 货币代码,
                c.currency_name as 货币名称,
                t.foreign_amount as 外币金额,
                t.local_amount as 本币金额,
                t.exchange_rate as 汇率,
                CASE
                    WHEN t.exchange_type = 'normal' THEN '普通兑换'
                    WHEN t.exchange_type = 'large_amount' THEN '大额兑换'
                    WHEN t.exchange_type = 'asset_mortgage' THEN '资产抵押'
                    ELSE ''
                END as 兑换类型,
                CASE
                    WHEN t.funding_source = 'salary' THEN '工资收入'
                    WHEN t.funding_source = 'business' THEN '经营所得'
                    WHEN t.funding_source = 'investment' THEN '投资收益'
                    WHEN t.funding_source = 'inheritance' THEN '继承所得'
                    WHEN t.funding_source = 'gift' THEN '赠与'
                    WHEN t.funding_source = 'loan' THEN '贷款'
                    WHEN t.funding_source = 'other' THEN '其他'
                    ELSE ''
                END as 资金来源
            FROM exchange_transactions t
            LEFT JOIN currencies c ON t.currency_id = c.id
            WHERE t.branch_id = :branch_id
                AND DATE(t.transaction_time) = :date
                AND t.direction = 'buy'
                AND t.bot_flag = 1
                AND t.status = 'completed'
            ORDER BY t.transaction_time ASC
        """)

        result = session.execute(sql, {'branch_id': branch_id, 'date': date_str})
        rows = result.fetchall()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"买入外币报表_{date_str}"

        # 设置标题
        headers = [
            '交易编号', '交易时间', '客户证件号', '货币代码', '货币名称',
            '外币金额', '本币金额(THB)', '汇率', '兑换类型', '资金来源'
        ]

        # 写入标题行
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

        # 写入数据行
        for row_num, row_data in enumerate(rows, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # 调整列宽
        column_widths = [20, 20, 20, 12, 15, 15, 18, 12, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        filename = f"BOT_BuyFX_{date_str}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error in export_buy_fx_excel: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'导出Excel失败: {str(e)}'
        }), 500

    finally:
        session.close()


@app_bot.route('/api/bot/export-sell-fx', methods=['GET'])
@token_required
@bot_permission_required('bot_report_export')
def export_sell_fx_excel():
    """
    导出卖出外币报表Excel

    GET /api/bot/export-sell-fx?date=2025-10-01

    查询参数:
    - date: 交易日期，默认昨天

    响应:
    Excel文件下载
    """
    session = SessionLocal()

    try:
        # 获取日期参数
        date_str = request.args.get('date')
        if not date_str:
            yesterday = datetime.now() - timedelta(days=1)
            date_str = yesterday.strftime('%Y-%m-%d')

        # 获取当前用户的branch_id
        branch_id = g.current_user.get('branch_id')

        # 查询数据
        sql = text("""
            SELECT
                t.transaction_no as 交易编号,
                DATE_FORMAT(t.transaction_time, '%Y-%m-%d %H:%i:%s') as 交易时间,
                t.customer_id as 客户证件号,
                c.currency_code as 货币代码,
                c.currency_name as 货币名称,
                t.foreign_amount as 外币金额,
                t.local_amount as 本币金额,
                t.exchange_rate as 汇率,
                CASE
                    WHEN t.exchange_type = 'normal' THEN '普通兑换'
                    WHEN t.exchange_type = 'large_amount' THEN '大额兑换'
                    WHEN t.exchange_type = 'asset_mortgage' THEN '资产抵押'
                    ELSE ''
                END as 兑换类型
            FROM exchange_transactions t
            LEFT JOIN currencies c ON t.currency_id = c.id
            WHERE t.branch_id = :branch_id
                AND DATE(t.transaction_time) = :date
                AND t.direction = 'sell'
                AND t.bot_flag = 1
                AND t.status = 'completed'
            ORDER BY t.transaction_time ASC
        """)

        result = session.execute(sql, {'branch_id': branch_id, 'date': date_str})
        rows = result.fetchall()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"卖出外币报表_{date_str}"

        # 设置标题
        headers = [
            '交易编号', '交易时间', '客户证件号', '货币代码', '货币名称',
            '外币金额', '本币金额(THB)', '汇率', '兑换类型'
        ]

        # 写入标题行
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

        # 写入数据行
        for row_num, row_data in enumerate(rows, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # 调整列宽
        column_widths = [20, 20, 20, 12, 15, 15, 18, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        filename = f"BOT_SellFX_{date_str}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error in export_sell_fx_excel: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'导出Excel失败: {str(e)}'
        }), 500

    finally:
        session.close()


@app_bot.route('/api/bot/save-buy-fx', methods=['POST'])
@token_required
def save_bot_buy_fx(current_user):
    """
    保存BOT买入外币报表记录

    POST /api/bot/save-buy-fx

    请求体:
    {
        "transaction_id": 12345,
        "report_date": "2025-10-01",
        "json_data": {...}
    }

    响应:
    {
        "success": true,
        "report_id": 1
    }
    """
    session = SessionLocal()

    try:
        request_data = request.get_json()

        if not request_data:
            return jsonify({
                'success': False,
                'message': '请求数据不能为空'
            }), 400

        transaction_id = request_data.get('transaction_id')
        report_date = request_data.get('report_date')
        json_data = request_data.get('json_data', {})

        if not transaction_id or not report_date:
            return jsonify({
                'success': False,
                'message': '缺少必要参数'
            }), 400

        # 插入记录
        sql = text("""
            INSERT INTO BOT_BuyFX (
                transaction_id, report_date, json_data,
                branch_id, operator_id, created_at
            ) VALUES (
                :transaction_id, :report_date, :json_data,
                :branch_id, :operator_id, NOW()
            )
        """)

        current_user = g.current_user
        params = {
            'transaction_id': transaction_id,
            'report_date': report_date,
            'json_data': json.dumps(json_data, ensure_ascii=False),
            'branch_id': current_user['branch_id'],
            'operator_id': current_user['id']
        }

        result = session.execute(sql, params)
        session.commit()

        return jsonify({
            'success': True,
            'report_id': result.lastrowid
        })

    except Exception as e:
        session.rollback()
        print(f"Error in save_bot_buy_fx: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'保存记录失败: {str(e)}'
        }), 500

    finally:
        session.close()


@app_bot.route('/api/bot/save-sell-fx', methods=['POST'])
@token_required
def save_bot_sell_fx(current_user):
    """
    保存BOT卖出外币报表记录

    POST /api/bot/save-sell-fx

    请求体:
    {
        "transaction_id": 12345,
        "report_date": "2025-10-01",
        "json_data": {...}
    }

    响应:
    {
        "success": true,
        "report_id": 1
    }
    """
    session = SessionLocal()

    try:
        request_data = request.get_json()

        if not request_data:
            return jsonify({
                'success': False,
                'message': '请求数据不能为空'
            }), 400

        transaction_id = request_data.get('transaction_id')
        report_date = request_data.get('report_date')
        json_data = request_data.get('json_data', {})

        if not transaction_id or not report_date:
            return jsonify({
                'success': False,
                'message': '缺少必要参数'
            }), 400

        # 插入记录
        sql = text("""
            INSERT INTO BOT_SellFX (
                transaction_id, report_date, json_data,
                branch_id, operator_id, created_at
            ) VALUES (
                :transaction_id, :report_date, :json_data,
                :branch_id, :operator_id, NOW()
            )
        """)

        current_user = g.current_user
        params = {
            'transaction_id': transaction_id,
            'report_date': report_date,
            'json_data': json.dumps(json_data, ensure_ascii=False),
            'branch_id': current_user['branch_id'],
            'operator_id': current_user['id']
        }

        result = session.execute(sql, params)
        session.commit()

        return jsonify({
            'success': True,
            'report_id': result.lastrowid
        })

    except Exception as e:
        session.rollback()
        print(f"Error in save_bot_sell_fx: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'保存记录失败: {str(e)}'
        }), 500

    finally:
        session.close()


# 错误处理
@app_bot.errorhandler(404)
def not_found(error):
    return jsonify({
        'success': False,
        'message': '接口不存在'
    }), 404


@app_bot.errorhandler(500)
def internal_error(error):
    return jsonify({
        'success': False,
        'message': '服务器内部错误'
    }), 500
