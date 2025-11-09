#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""分析BOT样本Excel格式"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.utils import get_column_letter

# Windows UTF-8
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def analyze_bot_excel():
    """分析BOT样本Excel"""
    
    file_path = r"D:\Code\ExchangeNew\Re\BOT( Save Excel, PDF, Bot).xlsx"
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return
    
    print("="*80)
    print("BOT样本Excel格式分析")
    print("="*80)
    
    wb = openpyxl.load_workbook(file_path)
    
    print(f"\n总共有 {len(wb.sheetnames)} 个工作表:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. {sheet_name}")
    
    # 分析每个工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"\n{'='*80}")
        print(f"工作表: {sheet_name}")
        print(f"{'='*80}")
        
        # 获取使用的区域
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"使用区域: {max_row} 行 x {max_col} 列")
        
        # 读取前10行内容
        print(f"\n前10行内容:")
        print("-"*80)
        
        for row_idx in range(1, min(11, max_row + 1)):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    row_data.append(f"{get_column_letter(col_idx)}: {value}")
            
            if row_data:
                print(f"第{row_idx}行: {' | '.join(row_data)}")
        
        # 分析列标题（假设在第1行或第2行）
        print(f"\n列标题分析:")
        print("-"*80)
        
        # 尝试第1行
        headers_row1 = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                headers_row1.append((col_idx, get_column_letter(col_idx), cell.value))
        
        if headers_row1:
            print("第1行列标题:")
            for col_num, col_letter, header in headers_row1:
                print(f"  列{col_letter}({col_num}): {header}")
        
        # 尝试第2行
        headers_row2 = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=2, column=col_idx)
            if cell.value:
                headers_row2.append((col_idx, get_column_letter(col_idx), cell.value))
        
        if headers_row2:
            print("\n第2行列标题:")
            for col_num, col_letter, header in headers_row2:
                print(f"  列{col_letter}({col_num}): {header}")
        
        # 分析数据行（从第3行开始）
        print(f"\n数据样本（第3-5行）:")
        print("-"*80)
        
        for row_idx in range(3, min(6, max_row + 1)):
            print(f"\n第{row_idx}行:")
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    col_letter = get_column_letter(col_idx)
                    # 获取对应的列标题
                    header = ""
                    if col_idx <= len(headers_row1):
                        header = headers_row1[col_idx-1][2] if headers_row1 else ""
                    print(f"  {col_letter} ({header}): {cell.value}")
    
    print(f"\n{'='*80}")
    print("分析完成")
    print(f"{'='*80}")

if __name__ == "__main__":
    analyze_bot_excel()

