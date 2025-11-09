# -*- coding: utf-8 -*-
"""
BOT报表生成服务 - 基于模板文件
复制标准模板，填充实际数据
版本: v2.0 (Template-based)
创建日期: 2025-10-08
"""

import os
import sys
import shutil
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import logging

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


class BOTTemplateBasedGenerator:
    """基于模板的BOT Excel报表生成器"""

    # 模板文件路径
    TEMPLATE_PATH = r"D:\Code\ExchangeNew\Re\BOT( Save Excel, PDF, Bot).xlsx"

    # 数据起始行号（根据模板分析）
    PROVIDER_INFO_DATA_START = 2  # Provider Info从第2行开始
    BUY_FX_DATA_START = 9  # Buy FX数据从第9行开始
    SELL_FX_DATA_START = 9  # Sell FX数据从第9行开始
    FCD_DATA_START = 8  # FCD数据从第8行开始

    @classmethod
    def generate_report(
        cls,
        db_session: Session,
        branch_id: int,
        report_month: int,  # 1-12
        report_year: int,  # 泰国佛历，如2568
        output_path: Optional[str] = None
    ) -> str:
        """
        生成BOT报表（保存到manager目录）

        Args:
            db_session: 数据库会话
            branch_id: 网点ID
            report_month: 报告月份 (1-12)
            report_year: 报告年份 (泰国佛历)
            output_path: 输出路径（可选，如不指定则保存到manager目录）

        Returns:
            生成的文件路径
        """
        try:
            # 检查模板文件
            if not os.path.exists(cls.TEMPLATE_PATH):
                raise FileNotFoundError(f"模板文件不存在: {cls.TEMPLATE_PATH}")

            # 复制模板到临时文件
            temp_path = cls._create_temp_copy()
            logger.info(f"已复制模板到: {temp_path}")

            # 打开工作簿
            wb = openpyxl.load_workbook(temp_path)

            # 1. 获取网点信息
            branch_info = cls._get_branch_info(db_session, branch_id)

            # 2. 填充Provider Info
            cls._fill_provider_info(wb, branch_info, report_month, report_year)

            # 3. 填充Buy FX数据
            cls._fill_buy_fx_data(wb, db_session, branch_id, report_month, report_year)

            # 4. 填充Sell FX数据
            cls._fill_sell_fx_data(wb, db_session, branch_id, report_month, report_year)

            # 5. 填充FCD数据
            cls._fill_fcd_data(wb, db_session, branch_id, report_month, report_year)

            # 确定保存路径（保存到manager目录）
            if not output_path:
                # 转换为公历年份
                gregorian_year = report_year - 543
                
                # 构建manager目录路径
                current_dir = os.path.dirname(os.path.abspath(__file__))
                manager_dir = os.path.join(current_dir, '..', 'manager', str(gregorian_year), f"{report_month:02d}")
                os.makedirs(manager_dir, exist_ok=True)
                
                # 生成文件名
                filename = f"BOT_Report_{gregorian_year}{report_month:02d}.xlsx"
                output_path = os.path.join(manager_dir, filename)
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"BOT报表已保存: {output_path}")
            
            return output_path

        except Exception as e:
            logger.error(f"生成BOT报表失败: {str(e)}")
            raise

        finally:
            # 清理临时文件
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass

    @classmethod
    def _create_temp_copy(cls) -> str:
        """创建模板的临时副本"""
        temp_dir = os.path.join(os.path.dirname(cls.TEMPLATE_PATH), 'temp')
        os.makedirs(temp_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        temp_filename = f'BOT_temp_{timestamp}.xlsx'
        temp_path = os.path.join(temp_dir, temp_filename)

        shutil.copy2(cls.TEMPLATE_PATH, temp_path)
        return temp_path

    @classmethod
    def _get_branch_info(cls, db_session: Session, branch_id: int) -> Dict[str, Any]:
        """获取网点信息"""
        sql = text("""
            SELECT
                branch_name,
                address,
                phone_number,
                license_number,
                branch_code,
                company_full_name,
                bot_sender_code,
                bot_branch_area_code,
                bot_license_number
            FROM branches
            WHERE id = :branch_id
        """)

        result = db_session.execute(sql, {'branch_id': branch_id}).fetchone()

        if not result:
            # 使用默认值
            return {
                'institution_code': '0105549142901',
                'license_holder_name': 'บริษัท วางสวิตต์ อินเตอร์เนชั่นแนล จำกัด',
                'license_no': 'MC325670003',
                'branch_name': 'บริษัท วางสวิตต์ อินเตอร์เนชั่นแนล จำกัด',
                'branch_area_code': '001'
            }

        branch_name, _, _, license_number, branch_code, company_full_name, bot_sender_code, bot_branch_area_code, bot_license_number = result

        institution_code = bot_sender_code or '0105549142901'
        license_holder_name = company_full_name or branch_name or 'บริษัท วางสวิตต์ อินเตอร์เนชั่นแนล จำกัด'
        license_no = bot_license_number or license_number or 'MC325670003'
        area_code = bot_branch_area_code or branch_code or str(branch_id).zfill(3)

        return {
            'institution_code': institution_code,
            'license_holder_name': license_holder_name,
            'license_no': license_no,
            'branch_name': branch_name or license_holder_name,
            'branch_area_code': area_code
        }

    @classmethod
    def _fill_provider_info(
        cls,
        wb: openpyxl.Workbook,
        branch_info: Dict[str, Any],
        report_month: int,
        report_year: int
    ):
        """填充Provider Info sheet"""
        ws = wb['Provider Info']

        # B2: 机构代码
        ws['B2'] = branch_info['institution_code']

        # B3: 许可证持有人名称
        ws['B3'] = branch_info['license_holder_name']

        # B4: License No
        ws['B4'] = branch_info['license_no']

        # B5: 分支机构名称
        ws['B5'] = branch_info['branch_name']

        # B6: 分支机构代码
        ws['B6'] = branch_info['branch_area_code']

        # B7: 月份名称（泰文）
        month_names_th = [
            '', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
            'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม'
        ]
        ws['B7'] = month_names_th[report_month] if 1 <= report_month <= 12 else ''

        # B8: 年份
        ws['B8'] = report_year

        # B9: 数据集日期（月末最后一天）
        # 计算该月最后一天
        if report_month == 12:
            next_month_year = report_year + 1
            next_month = 1
        else:
            next_month_year = report_year
            next_month = report_month + 1

        # 转换为公历（佛历 - 543）
        gregorian_year = report_year - 543

        import calendar
        last_day = calendar.monthrange(gregorian_year, report_month)[1]
        ws['B9'] = f"{gregorian_year}-{str(report_month).zfill(2)}-{str(last_day).zfill(2)}"

        logger.info(f"已填充Provider Info: {branch_info['branch_name']}, {report_year}/{report_month}")

    @classmethod
    def _fill_buy_fx_data(
        cls,
        wb: openpyxl.Workbook,
        db_session: Session,
        branch_id: int,
        report_month: int,
        report_year: int
    ):
        """填充Buy FX sheet数据"""
        ws = wb['Buy FX']

        # 转换为公历
        gregorian_year = report_year - 543

        # 查询数据
        sql = text("""
            SELECT
                transaction_date,
                transaction_no,
                customer_id_type,
                customer_id_number,
                customer_name,
                customer_country_code,
                buy_currency_code,
                buy_amount,
                local_amount,
                exchange_rate,
                usd_equivalent,
                remarks
            FROM bot_buyfx
            WHERE branch_id = :branch_id
              AND YEAR(transaction_date) = :year
              AND MONTH(transaction_date) = :month
            ORDER BY transaction_date, transaction_no
        """)

        results = db_session.execute(sql, {
            'branch_id': branch_id,
            'year': gregorian_year,
            'month': report_month
        }).fetchall()

        # 填充数据（从第9行开始）
        current_row = cls.BUY_FX_DATA_START
        for idx, row_data in enumerate(results, 1):
            # A列: 序号
            ws.cell(row=current_row, column=1, value=idx)

            # B-D列: 日期（年、月、日分开）
            tx_date = row_data[0]  # transaction_date
            ws.cell(row=current_row, column=2, value=tx_date.year + 543)  # 佛历年份
            ws.cell(row=current_row, column=3, value=tx_date.month)
            ws.cell(row=current_row, column=4, value=tx_date.day)

            # E列: 交易编号
            ws.cell(row=current_row, column=5, value=row_data[1])  # transaction_no

            # F-G列: 证件类型和号码
            ws.cell(row=current_row, column=6, value=cls._map_id_type(row_data[2]))
            ws.cell(row=current_row, column=7, value=row_data[3])  # customer_id_number

            # H列: 客户姓名
            ws.cell(row=current_row, column=8, value=row_data[4])  # customer_name

            # I-J列: 国籍代码
            country_code = row_data[5] or 'TH'
            ws.cell(row=current_row, column=9, value=cls._map_country_code(country_code))
            ws.cell(row=current_row, column=10, value=country_code)

            # K-L列: 货币代码
            currency_code = row_data[6] or 'USD'
            ws.cell(row=current_row, column=11, value=cls._map_currency_code(currency_code))
            ws.cell(row=current_row, column=12, value=currency_code)

            # M列: 外币金额
            ws.cell(row=current_row, column=13, value=float(row_data[7]) if row_data[7] else 0)

            # N列: 汇率
            ws.cell(row=current_row, column=14, value=float(row_data[9]) if row_data[9] else 0)

            # O-P列: 支付方式（默认现金）
            ws.cell(row=current_row, column=15, value='0753600001')
            ws.cell(row=current_row, column=16, value='เงินสด')  # 现金

            # Q列: 备注
            ws.cell(row=current_row, column=17, value=row_data[11] or '')

            current_row += 1

        logger.info(f"已填充Buy FX数据: {len(results)} 条记录")

    @classmethod
    def _fill_sell_fx_data(
        cls,
        wb: openpyxl.Workbook,
        db_session: Session,
        branch_id: int,
        report_month: int,
        report_year: int
    ):
        """填充Sell FX sheet数据"""
        ws = wb['Sell FX']

        gregorian_year = report_year - 543

        sql = text("""
            SELECT
                transaction_date,
                transaction_no,
                customer_id_type,
                customer_id_number,
                customer_name,
                customer_country_code,
                sell_currency_code,
                sell_amount,
                local_amount,
                exchange_rate,
                usd_equivalent,
                remarks
            FROM bot_sellfx
            WHERE branch_id = :branch_id
              AND YEAR(transaction_date) = :year
              AND MONTH(transaction_date) = :month
            ORDER BY transaction_date, transaction_no
        """)

        results = db_session.execute(sql, {
            'branch_id': branch_id,
            'year': gregorian_year,
            'month': report_month
        }).fetchall()

        current_row = cls.SELL_FX_DATA_START
        for idx, row_data in enumerate(results, 1):
            ws.cell(row=current_row, column=1, value=idx)

            tx_date = row_data[0]
            ws.cell(row=current_row, column=2, value=tx_date.year + 543)
            ws.cell(row=current_row, column=3, value=tx_date.month)
            ws.cell(row=current_row, column=4, value=tx_date.day)

            ws.cell(row=current_row, column=5, value=row_data[1])  # transaction_no
            ws.cell(row=current_row, column=6, value=cls._map_id_type(row_data[2]))
            ws.cell(row=current_row, column=7, value=row_data[3])
            ws.cell(row=current_row, column=8, value=row_data[4])

            country_code = row_data[5] or 'TH'
            ws.cell(row=current_row, column=9, value=cls._map_country_code(country_code))
            ws.cell(row=current_row, column=10, value=country_code)

            currency_code = row_data[6] or 'USD'
            ws.cell(row=current_row, column=11, value=cls._map_currency_code(currency_code))
            ws.cell(row=current_row, column=12, value=currency_code)

            ws.cell(row=current_row, column=13, value=float(row_data[7]) if row_data[7] else 0)
            ws.cell(row=current_row, column=14, value=float(row_data[9]) if row_data[9] else 0)

            ws.cell(row=current_row, column=15, value='0753600001')
            ws.cell(row=current_row, column=16, value='เงินสด')

            ws.cell(row=current_row, column=17, value=row_data[11] or '')

            current_row += 1

        logger.info(f"已填充Sell FX数据: {len(results)} 条记录")

    @classmethod
    def _fill_fcd_data(
        cls,
        wb: openpyxl.Workbook,
        db_session: Session,
        branch_id: int,
        report_month: int,
        report_year: int
    ):
        """填充FCD sheet数据"""
        ws = wb['FCD']

        gregorian_year = report_year - 543

        sql = text("""
            SELECT
                account_open_date,
                bank_name,
                account_number,
                currency_code,
                balance,
                transaction_amount,
                usd_equivalent,
                remarks
            FROM bot_fcd
            WHERE branch_id = :branch_id
              AND YEAR(account_open_date) = :year
              AND MONTH(account_open_date) = :month
            ORDER BY account_open_date, account_number
        """)

        results = db_session.execute(sql, {
            'branch_id': branch_id,
            'year': gregorian_year,
            'month': report_month
        }).fetchall()

        current_row = cls.FCD_DATA_START
        for idx, row_data in enumerate(results, 1):
            ws.cell(row=current_row, column=1, value=idx)

            open_date = row_data[0]
            ws.cell(row=current_row, column=2, value=open_date.year + 543)
            ws.cell(row=current_row, column=3, value=open_date.month)
            ws.cell(row=current_row, column=4, value=open_date.day)

            ws.cell(row=current_row, column=5, value=row_data[1])  # bank_name
            ws.cell(row=current_row, column=6, value=row_data[2])  # account_number
            ws.cell(row=current_row, column=7, value=row_data[3])  # currency_code
            ws.cell(row=current_row, column=8, value=float(row_data[4]) if row_data[4] else 0)  # balance
            ws.cell(row=current_row, column=9, value=float(row_data[5]) if row_data[5] else 0)  # transaction_amount
            ws.cell(row=current_row, column=10, value=row_data[7] or '')  # remarks

            current_row += 1

        logger.info(f"已填充FCD数据: {len(results)} 条记录")

    @staticmethod
    def _map_id_type(id_type: str) -> str:
        """映射证件类型到泰文代码"""
        mapping = {
            'Passport': '324002',
            'ID Card': '324001',
            'Thai ID': '324001',
            'National ID': '324001',
            'Corporate': '324004'
        }
        return mapping.get(id_type, '324002')  # 默认护照

    @staticmethod
    def _map_country_code(country_code: str) -> str:
        """映射国家代码到泰文代码"""
        # 这里简化处理，实际应该从Dimension sheet的lookup表获取
        mapping = {
            'TH': '176001',
            'CHN': '176003',
            'CN': '176003',
            'JP': '176004',
            'MY': '176005',
            'SG': '176006',
            'US': '176002',
            'GB': '176007'
        }
        return mapping.get(country_code.upper(), '176999')

    @staticmethod
    def _map_currency_code(currency_code: str) -> str:
        """映射货币代码到泰文代码"""
        mapping = {
            'USD': '0753500001',
            'EUR': '0753500002',
            'GBP': '0753500003',
            'JPY': '0753500004',
            'CNY': '0753500005',
            'SGD': '0753500006'
        }
        return mapping.get(currency_code.upper(), '0753500999')
