# -*- coding: utf-8 -*-
"""
BOT Excel多sheet生成服务
生成包含BuyFX、SellFX、FCD、Provider的多sheet Excel报表
版本: v1.0
创建日期: 2025-10-08
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
from sqlalchemy import text
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
import io

logger = logging.getLogger(__name__)


class BOTExcelService:
    """BOT多sheet Excel生成服务"""

    @staticmethod
    def generate_multi_sheet_excel(
        db_session: Session,
        branch_id: int,
        start_date: str,
        end_date: str,
        output_path: Optional[str] = None
    ) -> bytes:
        """
        生成BOT多sheet Excel报表

        Args:
            db_session: 数据库会话
            branch_id: 网点ID
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            output_path: 输出文件路径（可选，如果提供则保存到文件）

        Returns:
            Excel文件的字节流
        """
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()

            # 删除默认sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 1. 生成BuyFX sheet
            logger.info("生成BuyFX sheet...")
            BOTExcelService._create_buyfx_sheet(wb, db_session, branch_id, start_date, end_date)

            # 2. 生成SellFX sheet
            logger.info("生成SellFX sheet...")
            BOTExcelService._create_sellfx_sheet(wb, db_session, branch_id, start_date, end_date)

            # 3. 生成FCD sheet
            logger.info("生成FCD sheet...")
            BOTExcelService._create_fcd_sheet(wb, db_session, branch_id, start_date, end_date)

            # 4. 生成Provider sheet
            logger.info("生成Provider sheet...")
            BOTExcelService._create_provider_sheet(wb, db_session, branch_id, start_date, end_date)

            # 保存到内存或文件
            if output_path:
                wb.save(output_path)
                logger.info(f"BOT Excel已保存到: {output_path}")
                with open(output_path, 'rb') as f:
                    return f.read()
            else:
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return output.getvalue()

        except Exception as e:
            logger.error(f"生成BOT Excel失败: {str(e)}")
            raise

    @staticmethod
    def _create_buyfx_sheet(
        wb: openpyxl.Workbook,
        db_session: Session,
        branch_id: int,
        start_date: str,
        end_date: str
    ):
        """创建BuyFX sheet"""
        ws = wb.create_sheet("BOT_BuyFX")

        # 设置标题行
        headers = [
            '序号', '交易日期', '交易时间', '交易编号', '客户证件号', '客户姓名',
            '货币代码', '货币名称', '外币金额', '本币金额(THB)', '汇率',
            '兑换类型', '资金来源', '是否已上报', '上报时间'
        ]

        # 写入标题
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11, name='Arial')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=11)

        # 查询数据
        sql = text("""
            SELECT
                transaction_no,
                transaction_date,
                customer_id,
                customer_name,
                currency_code,
                currency_name,
                foreign_amount,
                local_amount_thb,
                exchange_rate,
                exchange_type,
                funding_source,
                is_reported,
                report_time,
                created_at
            FROM BOT_BuyFX
            WHERE branch_id = :branch_id
                AND DATE(created_at) BETWEEN :start_date AND :end_date
            ORDER BY created_at ASC
        """)

        result = db_session.execute(sql, {
            'branch_id': branch_id,
            'start_date': start_date,
            'end_date': end_date
        })

        # 写入数据
        row_num = 2
        for idx, row_data in enumerate(result, 1):
            # 解析交易时间
            tx_datetime = row_data[1]  # transaction_date
            tx_date = tx_datetime.strftime('%Y-%m-%d') if tx_datetime else ''
            tx_time = tx_datetime.strftime('%H:%M:%S') if tx_datetime else ''

            # 兑换类型翻译
            exchange_type_map = {
                'normal': '普通兑换',
                'large_amount': '大额兑换',
                'asset_mortgage': '资产抵押'
            }
            exchange_type_zh = exchange_type_map.get(row_data[9], row_data[9] or '')

            # 资金来源翻译
            funding_source_map = {
                'salary': '工资收入',
                'business': '经营所得',
                'investment': '投资收益',
                'inheritance': '继承所得',
                'gift': '赠与',
                'loan': '贷款',
                'other': '其他'
            }
            funding_source_zh = funding_source_map.get(row_data[10], row_data[10] or '')

            # 写入行数据
            data_row = [
                idx,  # 序号
                tx_date,  # 交易日期
                tx_time,  # 交易时间
                row_data[0],  # 交易编号
                row_data[2],  # 客户证件号
                row_data[3],  # 客户姓名
                row_data[4],  # 货币代码
                row_data[5],  # 货币名称
                float(row_data[6]) if row_data[6] else 0,  # 外币金额
                float(row_data[7]) if row_data[7] else 0,  # 本币金额
                float(row_data[8]) if row_data[8] else 0,  # 汇率
                exchange_type_zh,  # 兑换类型
                funding_source_zh,  # 资金来源
                '是' if row_data[11] else '否',  # 是否已上报
                row_data[12].strftime('%Y-%m-%d %H:%M:%S') if row_data[12] else ''  # 上报时间
            ]

            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 数字格式
                if col_num in [9, 10, 11]:
                    cell.number_format = '#,##0.00'

            row_num += 1

        # 调整列宽
        column_widths = [8, 12, 10, 18, 18, 15, 10, 12, 15, 18, 12, 12, 12, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # 冻结首行
        ws.freeze_panes = 'A2'

    @staticmethod
    def _create_sellfx_sheet(
        wb: openpyxl.Workbook,
        db_session: Session,
        branch_id: int,
        start_date: str,
        end_date: str
    ):
        """创建SellFX sheet"""
        ws = wb.create_sheet("BOT_SellFX")

        # 设置标题行
        headers = [
            '序号', '交易日期', '交易时间', '交易编号', '客户证件号', '客户姓名',
            '货币代码', '货币名称', '外币金额', '本币金额(THB)', '汇率',
            '兑换类型', '是否已上报', '上报时间'
        ]

        # 写入标题
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11, name='Arial')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=11)

        # 查询数据
        sql = text("""
            SELECT
                transaction_no,
                transaction_date,
                customer_id,
                customer_name,
                currency_code,
                currency_name,
                foreign_amount,
                local_amount_thb,
                exchange_rate,
                exchange_type,
                is_reported,
                report_time,
                created_at
            FROM BOT_SellFX
            WHERE branch_id = :branch_id
                AND DATE(created_at) BETWEEN :start_date AND :end_date
            ORDER BY created_at ASC
        """)

        result = db_session.execute(sql, {
            'branch_id': branch_id,
            'start_date': start_date,
            'end_date': end_date
        })

        # 写入数据
        row_num = 2
        for idx, row_data in enumerate(result, 1):
            tx_datetime = row_data[1]
            tx_date = tx_datetime.strftime('%Y-%m-%d') if tx_datetime else ''
            tx_time = tx_datetime.strftime('%H:%M:%S') if tx_datetime else ''

            exchange_type_map = {
                'normal': '普通兑换',
                'large_amount': '大额兑换',
                'asset_mortgage': '资产抵押'
            }
            exchange_type_zh = exchange_type_map.get(row_data[9], row_data[9] or '')

            data_row = [
                idx,
                tx_date,
                tx_time,
                row_data[0],
                row_data[2],
                row_data[3],
                row_data[4],
                row_data[5],
                float(row_data[6]) if row_data[6] else 0,
                float(row_data[7]) if row_data[7] else 0,
                float(row_data[8]) if row_data[8] else 0,
                exchange_type_zh,
                '是' if row_data[10] else '否',
                row_data[11].strftime('%Y-%m-%d %H:%M:%S') if row_data[11] else ''
            ]

            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_num in [9, 10, 11]:
                    cell.number_format = '#,##0.00'

            row_num += 1

        # 调整列宽
        column_widths = [8, 12, 10, 18, 18, 15, 10, 12, 15, 18, 12, 12, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.freeze_panes = 'A2'

    @staticmethod
    def _create_fcd_sheet(
        wb: openpyxl.Workbook,
        db_session: Session,
        branch_id: int,
        start_date: str,
        end_date: str
    ):
        """创建FCD sheet"""
        ws = wb.create_sheet("BOT_FCD")

        headers = [
            '序号', '交易日期', '交易时间', '交易编号', '客户证件号', '客户姓名',
            '货币代码', '货币名称', '交易方向', '外币金额', '本币金额(THB)', '汇率',
            '是否已上报', '上报时间'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11, name='Arial')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=11)

        sql = text("""
            SELECT
                transaction_no,
                transaction_date,
                customer_id,
                customer_name,
                currency_code,
                currency_name,
                transaction_direction,
                foreign_amount,
                local_amount_thb,
                exchange_rate,
                is_reported,
                report_time,
                created_at
            FROM BOT_FCD
            WHERE branch_id = :branch_id
                AND DATE(created_at) BETWEEN :start_date AND :end_date
            ORDER BY created_at ASC
        """)

        result = db_session.execute(sql, {
            'branch_id': branch_id,
            'start_date': start_date,
            'end_date': end_date
        })

        row_num = 2
        for idx, row_data in enumerate(result, 1):
            tx_datetime = row_data[1]
            tx_date = tx_datetime.strftime('%Y-%m-%d') if tx_datetime else ''
            tx_time = tx_datetime.strftime('%H:%M:%S') if tx_datetime else ''

            direction_map = {'buy': '买入', 'sell': '卖出'}
            direction_zh = direction_map.get(row_data[6], row_data[6] or '')

            data_row = [
                idx,
                tx_date,
                tx_time,
                row_data[0],
                row_data[2],
                row_data[3],
                row_data[4],
                row_data[5],
                direction_zh,
                float(row_data[7]) if row_data[7] else 0,
                float(row_data[8]) if row_data[8] else 0,
                float(row_data[9]) if row_data[9] else 0,
                '是' if row_data[10] else '否',
                row_data[11].strftime('%Y-%m-%d %H:%M:%S') if row_data[11] else ''
            ]

            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_num in [10, 11, 12]:
                    cell.number_format = '#,##0.00'

            row_num += 1

        column_widths = [8, 12, 10, 18, 18, 15, 10, 12, 10, 15, 18, 12, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.freeze_panes = 'A2'

    @staticmethod
    def _create_provider_sheet(
        wb: openpyxl.Workbook,
        db_session: Session,
        branch_id: int,
        start_date: str,
        end_date: str
    ):
        """创建Provider sheet"""
        ws = wb.create_sheet("BOT_Provider")

        headers = [
            '序号', '调节日期', '调节时间', '货币代码', '货币名称',
            '调节金额', '本币金额(THB)', '调节原因', '是否已上报', '上报时间'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11, name='Arial')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            cell.font = Font(bold=True, color='000000', size=11)

        sql = text("""
            SELECT
                adjustment_date,
                currency_code,
                currency_name,
                provider_amount,
                local_amount_thb,
                adjustment_reason,
                is_reported,
                report_time,
                created_at
            FROM BOT_Provider
            WHERE branch_id = :branch_id
                AND DATE(created_at) BETWEEN :start_date AND :end_date
            ORDER BY created_at ASC
        """)

        result = db_session.execute(sql, {
            'branch_id': branch_id,
            'start_date': start_date,
            'end_date': end_date
        })

        row_num = 2
        for idx, row_data in enumerate(result, 1):
            adj_datetime = row_data[0]
            adj_date = adj_datetime.strftime('%Y-%m-%d') if adj_datetime else ''
            adj_time = adj_datetime.strftime('%H:%M:%S') if adj_datetime else ''

            data_row = [
                idx,
                adj_date,
                adj_time,
                row_data[1],
                row_data[2],
                float(row_data[3]) if row_data[3] else 0,
                float(row_data[4]) if row_data[4] else 0,
                row_data[5] or '',
                '是' if row_data[6] else '否',
                row_data[7].strftime('%Y-%m-%d %H:%M:%S') if row_data[7] else ''
            ]

            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_num in [6, 7]:
                    cell.number_format = '#,##0.00'

            row_num += 1

        column_widths = [8, 12, 10, 10, 12, 15, 18, 30, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.freeze_panes = 'A2'
