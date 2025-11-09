#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完整的AMLO和BOT测试
1. 创建BOT表（如果不存在）
2. 生成3种AMLO报告（1-01, 1-02, 1-03），并填充完整数据
3. 生成BOT报表
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Windows UTF-8
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from services.db_service import DatabaseService
from models.exchange_models import ExchangeTransaction, Currency, Branch
from sqlalchemy import text
from datetime import datetime, date
import glob

def check_and_create_bot_tables():
    """检查并创建BOT表"""
    session = DatabaseService.get_session()
    
    try:
        # 检查表是否存在
        result = session.execute(text("SHOW TABLES LIKE 'BOT%'"))
        existing_tables = [row[0] for row in result]
        
        print("="*80)
        print("[1] 检查BOT表")
        print("="*80)
        print(f"现有BOT表: {existing_tables if existing_tables else '无'}")
        
        if 'BOT_BuyFX' in existing_tables and 'BOT_SellFX' in existing_tables:
            print("[OK] BOT表已存在\n")
            return True
        
        print("\n[INFO] 创建BOT表...")
        
        # 创建BOT_BuyFX表
        session.execute(text("""
            CREATE TABLE IF NOT EXISTS `BOT_BuyFX` (
                `id` INT AUTO_INCREMENT PRIMARY KEY COMMENT 'ID',
                `transaction_id` INT NOT NULL COMMENT '关联的交易流水ID',
                `transaction_no` VARCHAR(30) NOT NULL COMMENT '交易流水号',
                `transaction_date` DATE NOT NULL COMMENT '交易日期',
                `customer_id_type` VARCHAR(20) DEFAULT 'PASSPORT' COMMENT '证件类型',
                `customer_id_number` VARCHAR(50) NOT NULL COMMENT '证件号码',
                `customer_name` VARCHAR(100) NOT NULL COMMENT '客户姓名',
                `customer_country_code` VARCHAR(3) DEFAULT 'USA' COMMENT '客户国籍代码',
                `customer_country_name` VARCHAR(100) DEFAULT NULL COMMENT '客户国籍名称',
                `rate_type` VARCHAR(20) DEFAULT 'CASH' COMMENT '汇率类型',
                `buy_currency_code` VARCHAR(3) NOT NULL COMMENT '买入外币币种',
                `buy_amount` DECIMAL(15,2) NOT NULL COMMENT '买入外币金额',
                `local_currency_code` VARCHAR(3) NOT NULL DEFAULT 'THB' COMMENT '本币币种',
                `local_amount` DECIMAL(15,2) NOT NULL COMMENT '本币金额',
                `exchange_rate` DECIMAL(10,4) NOT NULL COMMENT '汇率',
                `usd_equivalent` DECIMAL(15,2) DEFAULT NULL COMMENT '美元等值',
                `remarks` TEXT DEFAULT NULL COMMENT '备注',
                `branch_id` INT NOT NULL COMMENT '网点ID',
                `operator_id` INT NOT NULL COMMENT '操作员ID',
                `bot_flag` INT DEFAULT 1 COMMENT 'BOT标记',
                `use_fcd` BOOLEAN DEFAULT FALSE COMMENT '是否使用FCD账户',
                `is_reported` BOOLEAN DEFAULT FALSE COMMENT '是否已上报',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
                KEY `idx_transaction_id` (`transaction_id`),
                KEY `idx_transaction_date` (`transaction_date`),
                KEY `idx_customer_id` (`customer_id_number`),
                KEY `idx_branch_id` (`branch_id`),
                KEY `idx_bot_flag` (`bot_flag`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='BOT买入外币报表数据'
        """))
        
        # 创建BOT_SellFX表
        session.execute(text("""
            CREATE TABLE IF NOT EXISTS `BOT_SellFX` (
                `id` INT AUTO_INCREMENT PRIMARY KEY COMMENT 'ID',
                `transaction_id` INT NOT NULL COMMENT '关联的交易流水ID',
                `transaction_no` VARCHAR(30) NOT NULL COMMENT '交易流水号',
                `transaction_date` DATE NOT NULL COMMENT '交易日期',
                `customer_id_type` VARCHAR(20) DEFAULT 'PASSPORT' COMMENT '证件类型',
                `customer_id_number` VARCHAR(50) NOT NULL COMMENT '证件号码',
                `customer_name` VARCHAR(100) NOT NULL COMMENT '客户姓名',
                `customer_country_code` VARCHAR(3) DEFAULT 'USA' COMMENT '客户国籍代码',
                `customer_country_name` VARCHAR(100) DEFAULT NULL COMMENT '客户国籍名称',
                `rate_type` VARCHAR(20) DEFAULT 'CASH' COMMENT '汇率类型',
                `sell_currency_code` VARCHAR(3) NOT NULL COMMENT '卖出外币币种',
                `sell_amount` DECIMAL(15,2) NOT NULL COMMENT '卖出外币金额',
                `local_currency_code` VARCHAR(3) NOT NULL DEFAULT 'THB' COMMENT '本币币种',
                `local_amount` DECIMAL(15,2) NOT NULL COMMENT '本币金额',
                `exchange_rate` DECIMAL(10,4) NOT NULL COMMENT '汇率',
                `usd_equivalent` DECIMAL(15,2) DEFAULT NULL COMMENT '美元等值',
                `remarks` TEXT DEFAULT NULL COMMENT '备注',
                `branch_id` INT NOT NULL COMMENT '网点ID',
                `operator_id` INT NOT NULL COMMENT '操作员ID',
                `bot_flag` INT DEFAULT 1 COMMENT 'BOT标记',
                `use_fcd` BOOLEAN DEFAULT FALSE COMMENT '是否使用FCD账户',
                `is_reported` BOOLEAN DEFAULT FALSE COMMENT '是否已上报',
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
                KEY `idx_transaction_id` (`transaction_id`),
                KEY `idx_transaction_date` (`transaction_date`),
                KEY `idx_customer_id` (`customer_id_number`),
                KEY `idx_branch_id` (`branch_id`),
                KEY `idx_bot_flag` (`bot_flag`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='BOT卖出外币报表数据'
        """))
        
        session.commit()
        
        print("[OK] BOT表创建成功")
        print("  - BOT_BuyFX")
        print("  - BOT_SellFX\n")
        
        return True
        
    except Exception as e:
        session.rollback()
        print(f"[ERROR] 创建BOT表失败: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        DatabaseService.close_session(session)

def create_transaction_with_full_data(branch_id, customer_data, transaction_type='buy'):
    """创建包含完整数据的交易"""
    session = DatabaseService.get_session()
    
    try:
        # 获取USD币种
        usd = session.query(Currency).filter_by(currency_code='USD').first()
        if not usd:
            print("[ERROR] USD币种不存在")
            return None
        
        # 生成交易号
        from utils.transaction_utils import generate_transaction_no
        tx_no = generate_transaction_no(branch_id, session)
        
        # 计算本币金额
        amount = customer_data['amount']
        rate = customer_data['rate']
        local_amount = amount * rate
        
        # 创建交易记录
        now = datetime.now()
        transaction = ExchangeTransaction(
            transaction_no=tx_no,
            branch_id=branch_id,
            currency_id=usd.id,
            type=transaction_type,
            amount=amount,
            rate=rate,
            local_amount=local_amount,
            customer_name=customer_data['name'],
            customer_id=customer_data['id'],
            customer_country_code=customer_data.get('country_code', 'USA'),
            customer_address=customer_data.get('address'),
            occupation=customer_data.get('occupation'),
            transaction_date=date.today(),
            transaction_time=now.strftime('%H:%M:%S'),
            status='completed',
            created_at=now,
            operator_id=1,
            seqno=1
        )
        
        session.add(transaction)
        session.flush()
        
        tx_id = transaction.id
        session.commit()
        
        print(f"[OK] 交易创建成功: {tx_no}")
        print(f"  ID: {tx_id} | {amount:,} USD = {local_amount:,} THB")
        
        return {
            'id': tx_id,
            'transaction_no': tx_no,
            'currency_id': usd.id,
            'amount': amount,
            'local_amount': local_amount,
            'rate': rate,
            'type': transaction_type
        }
        
    except Exception as e:
        session.rollback()
        print(f"[ERROR] 创建交易失败: {e}")
        return None
    finally:
        DatabaseService.close_session(session)

def generate_amlo_report_with_full_data(branch_id, transaction_data, customer_data, report_format):
    """生成包含完整数据的AMLO报告"""
    session = DatabaseService.get_session()
    
    try:
        from services.repform.report_data_service import ReportDataService
        from services.pdf.amlo_pdf_generator import AMLOPDFGenerator
        
        # 准备预约数据
        reservation_data = {
            'report_type': report_format,
            'customer_id': customer_data['id'],
            'customer_name': customer_data['name'],
            'currency_id': transaction_data['currency_id'],
            'direction': transaction_data['type'],
            'amount': transaction_data['amount'],
            'local_amount': transaction_data['local_amount'],
            'rate': transaction_data['rate'],
            'trigger_type': '1',
            'transaction_id': transaction_data['id'],
            'branch_id': branch_id,
            'operator_id': 1,
            'form_data': {
                'occupation': customer_data.get('occupation', '未知'),
                'address': customer_data.get('address', '未知'),
                'phone': customer_data.get('phone', '未知'),
                'purpose': customer_data.get('purpose', '未知'),
                'funding_source': customer_data.get('funding_source', '未知'),
                'remarks': customer_data.get('remarks', '')
            },
            'exchange_type': 'normal'
        }
        
        # 保存预约记录
        reservation_id = ReportDataService.save_reservation(session, reservation_data)
        
        # 映射report_type
        report_type_map = {
            'AMLO-1-01': 'CTR',
            'AMLO-1-02': 'ATR',
            'AMLO-1-03': 'STR'
        }
        
        # 准备报告数据
        report_data = {
            'report_type': report_type_map.get(report_format, 'CTR'),
            'report_format': report_format,
            'branch_id': branch_id,
            'reserved_id': reservation_id,
            'transaction_id': transaction_data['id'],
            'customer_id': customer_data['id'],
            'customer_name': customer_data['name'],
            'transaction_amount': transaction_data['local_amount'],
            'transaction_date': date.today(),
            'operator_id': 1,
            'language': 'zh',
            'pdf_filename': '',
            'pdf_path': ''
        }
        
        # 保存AMLO报告记录
        report_id = ReportDataService.save_amlo_report(session, report_data)
        
        # 获取报告记录
        report_record = session.execute(text("""
            SELECT * FROM AMLOReport WHERE id = :id
        """), {'id': report_id}).fetchone()
        
        # 准备PDF数据 - 从预约表单获取完整信息
        reservation = session.execute(text("""
            SELECT * FROM Reserved_Transaction WHERE id = :id
        """), {'id': reservation_id}).fetchone()
        
        # 解析form_data
        import json
        form_data = json.loads(reservation.form_data) if reservation.form_data else {}
        
        pdf_data = {
            'report_no': report_record.report_no,
            'report_date': datetime.now().strftime('%d/%m/%Y'),
            'customer_name': customer_data['name'],
            'customer_id': customer_data['id'],
            'customer_id_type': customer_data.get('id_type', 'Passport'),
            'occupation': form_data.get('occupation', customer_data.get('occupation', '')),
            'address': form_data.get('address', customer_data.get('address', '')),
            'phone': form_data.get('phone', customer_data.get('phone', '')),
            'nationality': customer_data.get('country_code', 'USA'),
            'transaction_date': date.today().strftime('%d/%m/%Y'),
            'transaction_type': '买入外币' if transaction_data['type'] == 'buy' else '卖出外币',
            'currency': 'USD',
            'foreign_amount': float(transaction_data['amount']),
            'exchange_rate': float(transaction_data['rate']),
            'thb_amount': float(transaction_data['local_amount']),
            'purpose': form_data.get('purpose', customer_data.get('purpose', '')),
            'funding_source': form_data.get('funding_source', customer_data.get('funding_source', '')),
            'remarks': form_data.get('remarks', customer_data.get('remarks', ''))
        }
        
        # 确定输出路径
        current_dir = os.path.dirname(os.path.abspath(__file__))
        year = date.today().year
        month = date.today().month
        
        manager_dir = os.path.join(current_dir, '..', 'manager', str(year), f"{month:02d}")
        os.makedirs(manager_dir, exist_ok=True)
        
        pdf_filename = f"{report_format}_R{report_record.report_no}.pdf"
        pdf_path = os.path.join(manager_dir, pdf_filename)
        
        # 生成PDF
        generator = AMLOPDFGenerator()
        result_path = generator.generate_pdf(report_format, pdf_data, pdf_path)
        
        # 更新数据库中的PDF路径
        session.execute(text("""
            UPDATE AMLOReport 
            SET pdf_filename = :filename, pdf_path = :path
            WHERE id = :id
        """), {
            'filename': pdf_filename,
            'path': pdf_path,
            'id': report_id
        })
        
        session.commit()
        
        file_size = os.path.getsize(pdf_path) if os.path.exists(pdf_path) else 0
        
        print(f"[OK] {report_format}报告生成成功")
        print(f"  报告ID: {report_id} | 文件: {pdf_filename} | 大小: {file_size/1024:.2f} KB")
        
        return {
            'report_id': report_id,
            'pdf_path': pdf_path,
            'pdf_filename': pdf_filename
        }
        
    except Exception as e:
        session.rollback()
        print(f"[ERROR] 生成{report_format}报告失败: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        DatabaseService.close_session(session)

def create_bot_record(branch_id, transaction_data, customer_data):
    """创建BOT记录"""
    session = DatabaseService.get_session()
    
    try:
        # 计算美元等值
        usd_equiv = transaction_data['amount']  # 本身就是USD
        
        sql = text("""
            INSERT INTO BOT_BuyFX (
                transaction_id, transaction_no, transaction_date,
                customer_id_type, customer_id_number, customer_name,
                customer_country_code, customer_country_name,
                rate_type, buy_currency_code, buy_amount,
                local_currency_code, local_amount, exchange_rate,
                usd_equivalent, remarks,
                branch_id, operator_id, bot_flag, use_fcd,
                created_at
            ) VALUES (
                :tx_id, :tx_no, :tx_date,
                :id_type, :id_number, :name,
                :country_code, :country_name,
                :rate_type, :currency, :amount,
                'THB', :local_amount, :rate,
                :usd_equiv, :remarks,
                :branch_id, :operator_id, 1, FALSE,
                NOW()
            )
        """)
        
        result = session.execute(sql, {
            'tx_id': transaction_data['id'],
            'tx_no': transaction_data['transaction_no'],
            'tx_date': date.today(),
            'id_type': customer_data.get('id_type', 'PASSPORT'),
            'id_number': customer_data['id'],
            'name': customer_data['name'],
            'country_code': customer_data.get('country_code', 'USA'),
            'country_name': customer_data.get('country_name', 'United States'),
            'rate_type': 'CASH',
            'currency': 'USD',
            'amount': transaction_data['amount'],
            'local_amount': transaction_data['local_amount'],
            'rate': transaction_data['rate'],
            'usd_equiv': usd_equiv,
            'remarks': f"测试BOT记录 - {customer_data.get('remarks', '')}",
            'branch_id': branch_id,
            'operator_id': 1
        })
        
        session.commit()
        bot_id = result.lastrowid
        
        print(f"[OK] BOT记录创建成功: ID={bot_id}")
        
        return bot_id
        
    except Exception as e:
        session.rollback()
        print(f"[ERROR] 创建BOT记录失败: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        DatabaseService.close_session(session)

def generate_bot_excel():
    """生成BOT Excel报表"""
    session = DatabaseService.get_session()
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # 查询今天的BOT数据
        query = text("""
            SELECT 
                b.*,
                br.branch_name
            FROM BOT_BuyFX b
            LEFT JOIN branches br ON b.branch_id = br.id
            WHERE b.transaction_date = :today
            ORDER BY b.created_at DESC
        """)
        
        result = session.execute(query, {'today': date.today()})
        records = result.fetchall()
        
        if not records:
            print("[WARN] 没有找到今天的BOT记录")
            return None
        
        # 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOT BuyFX Report"
        
        # 标题行
        headers = [
            "ID", "交易号", "交易日期", "证件类型", "证件号码",
            "客户姓名", "国籍", "币种", "外币金额", "汇率",
            "泰铢金额", "美元等值", "备注", "网点"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=11, name='Arial')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        for row_idx, record in enumerate(records, 2):
            data = [
                record.id,
                record.transaction_no,
                record.transaction_date.strftime('%Y-%m-%d'),
                record.customer_id_type,
                record.customer_id_number,
                record.customer_name,
                record.customer_country_code,
                record.buy_currency_code,
                float(record.buy_amount),
                float(record.exchange_rate),
                float(record.local_amount),
                float(record.usd_equivalent) if record.usd_equivalent else 0,
                record.remarks or '',
                record.branch_name or 'N/A'
            ]
            
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 保存文件
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        filename = f"BOT_BuyFX_{date.today().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(exports_dir, filename)
        
        wb.save(filepath)
        
        file_size = os.path.getsize(filepath)
        print(f"\n[OK] BOT Excel生成成功")
        print(f"  文件: {filename} | 大小: {file_size/1024:.2f} KB | 记录数: {len(records)}")
        print(f"  路径: {os.path.abspath(filepath)}")
        
        return filepath
        
    except Exception as e:
        print(f"[ERROR] 生成BOT Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        DatabaseService.close_session(session)

def main():
    print("="*80)
    print("完整的AMLO和BOT合规系统测试")
    print("="*80)
    print(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # 步骤1: 检查并创建BOT表
    if not check_and_create_bot_tables():
        print("\n[ERROR] 测试终止: BOT表创建失败")
        return
    
    # 获取A005网点
    session = DatabaseService.get_session()
    try:
        branch = session.query(Branch).filter_by(branch_code='A005').first()
        if not branch:
            branch = session.query(Branch).filter_by(id=1).first()
        branch_id = branch.id
        print(f"[OK] 测试网点: {branch.branch_name} (Code: {branch.branch_code})\n")
    finally:
        DatabaseService.close_session(session)
    
    # 步骤2: 生成3种AMLO报告
    print("="*80)
    print("[2] 生成AMLO报告（1-01, 1-02, 1-03）")
    print("="*80)
    
    test_cases = [
        {
            'report_format': 'AMLO-1-01',
            'customer': {
                'id': 'TEST_AMLO_101',
                'name': '张三（现金交易）',
                'id_type': 'PASSPORT',
                'country_code': 'CHN',
                'country_name': 'China',
                'address': '北京市朝阳区建国路100号',
                'occupation': '商人',
                'phone': '0861234567890',
                'purpose': '旅游支出',
                'funding_source': '个人储蓄',
                'remarks': '测试AMLO-1-01现金交易报告',
                'amount': 60000,
                'rate': 35.50
            }
        },
        {
            'report_format': 'AMLO-1-02',
            'customer': {
                'id': 'TEST_AMLO_102',
                'name': '李四（资产交易）',
                'id_type': 'PASSPORT',
                'country_code': 'USA',
                'country_name': 'United States',
                'address': 'New York, 5th Avenue 123',
                'occupation': '企业家',
                'phone': '0861987654321',
                'purpose': '房产投资',
                'funding_source': '公司收入',
                'remarks': '测试AMLO-1-02资产交易报告',
                'amount': 250000,
                'rate': 35.60
            }
        },
        {
            'report_format': 'AMLO-1-03',
            'customer': {
                'id': 'TEST_AMLO_103',
                'name': '王五（可疑交易）',
                'id_type': 'ID_CARD',
                'country_code': 'THA',
                'country_name': 'Thailand',
                'address': '曼谷市素坤逸路88号',
                'occupation': '自由职业者',
                'phone': '0821234567',
                'purpose': '不明',
                'funding_source': '现金',
                'remarks': '测试AMLO-1-03可疑交易报告 - 客户行为异常',
                'amount': 45000,
                'rate': 35.45
            }
        }
    ]
    
    amlo_results = []
    bot_transactions = []
    
    for idx, test_case in enumerate(test_cases, 1):
        print(f"\n[{idx}] 生成{test_case['report_format']}报告...")
        print("-"*80)
        
        # 创建交易
        transaction_data = create_transaction_with_full_data(
            branch_id,
            test_case['customer'],
            'buy'
        )
        
        if not transaction_data:
            print(f"[ERROR] 跳过{test_case['report_format']}\n")
            continue
        
        # 生成AMLO报告
        amlo_result = generate_amlo_report_with_full_data(
            branch_id,
            transaction_data,
            test_case['customer'],
            test_case['report_format']
        )
        
        if amlo_result:
            amlo_results.append(amlo_result)
            bot_transactions.append((transaction_data, test_case['customer']))
        
        print()
    
    # 步骤3: 创建BOT记录
    print("="*80)
    print("[3] 创建BOT记录")
    print("="*80)
    
    bot_ids = []
    for transaction_data, customer_data in bot_transactions:
        bot_id = create_bot_record(branch_id, transaction_data, customer_data)
        if bot_id:
            bot_ids.append(bot_id)
    
    # 步骤4: 生成BOT Excel
    print("\n" + "="*80)
    print("[4] 生成BOT Excel报表")
    print("="*80)
    
    bot_excel_path = generate_bot_excel()
    
    # 最终总结
    print("\n" + "="*80)
    print("测试完成总结")
    print("="*80)
    
    print(f"\n✅ AMLO报告: 生成 {len(amlo_results)}/3 个")
    for result in amlo_results:
        print(f"  - {result['pdf_filename']}")
    
    print(f"\n✅ BOT记录: 创建 {len(bot_ids)} 条")
    print(f"  - 记录ID: {bot_ids}")
    
    if bot_excel_path:
        print(f"\n✅ BOT Excel: {os.path.basename(bot_excel_path)}")
    
    print(f"\n📂 查看文件:")
    if amlo_results:
        manager_dir = os.path.dirname(os.path.abspath(amlo_results[0]['pdf_path']))
        print(f"  AMLO: explorer \"{manager_dir}\"")
    
    if bot_excel_path:
        exports_dir = os.path.dirname(os.path.abspath(bot_excel_path))
        print(f"  BOT:  explorer \"{exports_dir}\"")
    
    print(f"\n🌐 查看报告:")
    print(f"  - AMLO: http://localhost:8080/amlo/reports")
    print(f"  - BOT:  http://localhost:8080/bot/reports")

if __name__ == "__main__":
    main()

