#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
端到端BOT测试 - 真实生成Excel报表文件
直接调用后端服务生成BOT报表
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Windows UTF-8
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from services.db_service import DatabaseService
from models.exchange_models import ExchangeTransaction, Currency, Branch, BOT_BuyFX, BOT_SellFX
from sqlalchemy import text
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def create_bot_buyfx_record(branch_id, transaction_data):
    """创建BOT BuyFX记录"""
    session = DatabaseService.get_session()
    
    try:
        # 创建BOT_BuyFX记录
        bot_record = BOT_BuyFX(
            transaction_id=transaction_data['id'],
            transaction_date=date.today(),
            branch_id=branch_id,
            currency_code='USD',
            foreign_amount=transaction_data['amount'],
            exchange_rate=transaction_data['rate'],
            thb_amount=transaction_data['local_amount'],
            customer_type='Individual',
            purpose_code='TRV',  # Travel
            remarks='端到端测试BOT BuyFX记录',
            created_at=datetime.now()
        )
        
        session.add(bot_record)
        session.commit()
        
        bot_id = bot_record.id
        
        print(f"[OK] BOT BuyFX记录创建成功")
        print(f"  记录ID: {bot_id}")
        print(f"  币种: USD")
        print(f"  金额: {transaction_data['amount']:,} USD = {transaction_data['local_amount']:,} THB")
        
        return bot_id
        
    except Exception as e:
        session.rollback()
        print(f"[ERROR] 创建BOT记录失败: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        DatabaseService.close_session(session)

def generate_bot_excel(branch_id, report_type='BuyFX'):
    """生成BOT Excel报表"""
    session = DatabaseService.get_session()
    
    try:
        # 查询数据
        if report_type == 'BuyFX':
            query = text("""
                SELECT 
                    b.id,
                    b.transaction_date,
                    b.currency_code,
                    b.foreign_amount,
                    b.exchange_rate,
                    b.thb_amount,
                    b.customer_type,
                    b.purpose_code,
                    br.branch_name
                FROM BOT_BuyFX b
                LEFT JOIN branches br ON b.branch_id = br.id
                WHERE b.branch_id = :branch_id
                AND b.transaction_date = :today
                ORDER BY b.created_at DESC
                LIMIT 10
            """)
            
            result = session.execute(query, {
                'branch_id': branch_id,
                'today': date.today()
            })
            
            records = result.fetchall()
            
            if not records:
                print("[WARN] 没有找到BOT BuyFX记录")
                return None
            
            print(f"\n[OK] 找到 {len(records)} 条BOT BuyFX记录")
            
            # 创建Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOT BuyFX"
            
            # 标题行
            headers = [
                "ID", "交易日期", "币种", "外币金额", "汇率", 
                "泰铢金额", "客户类型", "用途代码", "网点"
            ]
            
            # 写入标题
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # 写入数据
            for row_idx, record in enumerate(records, 2):
                data = [
                    record.id,
                    record.transaction_date.strftime('%Y-%m-%d'),
                    record.currency_code,
                    float(record.foreign_amount),
                    float(record.exchange_rate),
                    float(record.thb_amount),
                    record.customer_type,
                    record.purpose_code,
                    record.branch_name or 'N/A'
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left' if isinstance(value, str) else 'right')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 15
            
            # 保存文件
            exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
            os.makedirs(exports_dir, exist_ok=True)
            
            filename = f"BOT_BuyFX_{date.today().strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(exports_dir, filename)
            
            wb.save(filepath)
            
            print(f"\n[OK] BOT Excel生成成功!")
            print(f"  文件名: {filename}")
            print(f"  完整路径: {os.path.abspath(filepath)}")
            print(f"  文件大小: {os.path.getsize(filepath)} 字节")
            print(f"  记录数: {len(records)}")
            
            return filepath
            
    except Exception as e:
        print(f"[ERROR] 生成BOT Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        DatabaseService.close_session(session)

def verify_files(amlo_path, bot_path):
    """验证所有文件存在"""
    print("\n" + "="*80)
    print("[最终验证] 文件存在性检查")
    print("="*80)
    
    files_ok = True
    
    # 验证AMLO PDF
    if os.path.exists(amlo_path):
        print(f"\n[OK] AMLO PDF存在")
        print(f"  路径: {os.path.abspath(amlo_path)}")
        print(f"  大小: {os.path.getsize(amlo_path) / 1024:.2f} KB")
    else:
        print(f"\n[ERROR] AMLO PDF不存在: {amlo_path}")
        files_ok = False
    
    # 验证BOT Excel
    if bot_path and os.path.exists(bot_path):
        print(f"\n[OK] BOT Excel存在")
        print(f"  路径: {os.path.abspath(bot_path)}")
        print(f"  大小: {os.path.getsize(bot_path) / 1024:.2f} KB")
    elif bot_path:
        print(f"\n[ERROR] BOT Excel不存在: {bot_path}")
        files_ok = False
    
    return files_ok

def main():
    print("="*80)
    print("端到端AMLO & BOT测试 - 完整文件生成验证")
    print("="*80)
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # 步骤1: 运行AMLO测试（生成交易和AMLO报告）
    print("="*80)
    print("[步骤1] 运行AMLO测试")
    print("="*80)
    
    # 导入AMLO测试模块
    import end_to_end_amlo_test
    from end_to_end_amlo_test import create_real_transaction, create_reservation_and_report
    
    # 获取A005网点
    session = DatabaseService.get_session()
    try:
        branch = session.query(Branch).filter_by(branch_code='A005').first()
        if not branch:
            print("[WARN] A005网点不存在，使用Branch 1")
            branch = session.query(Branch).filter_by(id=1).first()
        
        branch_id = branch.id
        print(f"[OK] 使用网点: {branch.branch_name} (ID: {branch_id}, Code: {branch.branch_code})\n")
    finally:
        DatabaseService.close_session(session)
    
    # 创建交易
    transaction_data = create_real_transaction(
        branch_id=branch_id,
        customer_id='TEST_FULL_E2E',
        customer_name='完整测试客户',
        amount_usd=80000,
        rate=35.60
    )
    
    if not transaction_data:
        print("\n[ERROR] 测试终止: AMLO交易创建失败")
        return
    
    # 创建AMLO报告
    amlo_report = create_reservation_and_report(
        branch_id=branch_id,
        transaction_data=transaction_data,
        customer_id='TEST_FULL_E2E',
        customer_name='完整测试客户'
    )
    
    if not amlo_report:
        print("\n[ERROR] 测试终止: AMLO报告创建失败")
        return
    
    amlo_pdf_path = amlo_report['pdf_path']
    
    # 步骤2: 创建BOT记录
    print("\n" + "="*80)
    print("[步骤2] 创建BOT BuyFX记录")
    print("="*80)
    
    bot_id = create_bot_buyfx_record(branch_id, transaction_data)
    
    if not bot_id:
        print("\n[ERROR] BOT记录创建失败")
    
    # 步骤3: 生成BOT Excel
    print("\n" + "="*80)
    print("[步骤3] 生成BOT Excel报表")
    print("="*80)
    
    bot_excel_path = generate_bot_excel(branch_id, 'BuyFX')
    
    # 步骤4: 验证所有文件
    all_ok = verify_files(amlo_pdf_path, bot_excel_path)
    
    # 最终总结
    print("\n" + "="*80)
    print("完整测试总结")
    print("="*80)
    
    if all_ok:
        print("\n[OK] 所有文件生成成功!")
    else:
        print("\n[WARN] 部分文件生成失败")
    
    print(f"\n生成的文件:")
    print(f"  1. AMLO报告PDF: {os.path.abspath(amlo_pdf_path)}")
    if bot_excel_path:
        print(f"  2. BOT报表Excel: {os.path.abspath(bot_excel_path)}")
    
    print(f"\n数据库记录:")
    print(f"  - 交易ID: {transaction_data['id']}")
    print(f"  - AMLO报告ID: {amlo_report['report_id']}")
    if bot_id:
        print(f"  - BOT记录ID: {bot_id}")
    
    print(f"\n打开文件目录:")
    print(f"  AMLO: explorer \"{os.path.dirname(os.path.abspath(amlo_pdf_path))}\"")
    if bot_excel_path:
        print(f"  BOT:  explorer \"{os.path.dirname(os.path.abspath(bot_excel_path))}\"")
    
    print(f"\n查看报告:")
    print(f"  - AMLO报告列表: http://localhost:8080/amlo/reports")
    print(f"  - BOT报表查询: http://localhost:8080/bot/reports")

if __name__ == "__main__":
    main()

